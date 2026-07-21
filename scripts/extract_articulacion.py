"""
Regenera js/articulacion.js a partir de data/articulacion_asig_comp.xlsx.

Uso:
    python scripts/extract_articulacion.py

Requiere: openpyxl  (pip install openpyxl)

Lee la matriz de articulación asignatura-competencia (una hoja por carrera/
modalidad) y escribe un único archivo JS con "const ARTICULACION = {...};"
que el dashboard carga directamente (sin fetch, para que funcione abriendo
index.html sin necesidad de un servidor local).

Los códigos de competencia de esa matriz (RAC1..RAC8, a veces con CTI1..CTI4
intercalados) NO se corresponden por posición con los códigos CE/CT que usa
el resto del dashboard -- el orden varía por carrera. Por eso cada columna
se empareja con un código CE/CT de Cuadros_oficiales_por_carrera.xlsx
comparando el texto completo de la descripción (normalizado). Cuando una
columna no tiene ningún código CE/CT candidato con el que coincida su
descripción, se conserva la etiqueta nativa de la hoja (p.ej. "RAC3") en vez
de inventar un número.
"""
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLS = ROOT / "data" / "articulacion_asig_comp.xlsx"
CUADROS = ROOT / "data" / "Cuadros_oficiales_por_carrera.xlsx"
OUT = ROOT / "js" / "articulacion.js"

COMP_ORDER = ["CE1", "CE2", "CE3", "CE4", "CT1", "CT2", "CT3", "CT4"]

# hoja de articulacion_asig_comp.xlsx -> (hoja de Cuadros_oficiales_por_carrera.xlsx, (carrera, modalidad) tal como
# aparecen en js/data.js / programs de app.js)
SHEET_MAP = {
    "CONTABILIDAD ": ("Contabilidad", ("Contabilidad y Auditoría", "Presencial")),
    "ADMINISTRACION DE EMPRESAS": ("Administración", ("Administración", "Presencial")),
    "ECONOMIA PRESENCIAL": ("Economía Presencial", ("Economía", "Presencial")),
    "ECONOMIA EN LINEA": ("Economía en Línea", ("Economía", "En línea")),
    "TURISMO PRESENCIAL": ("Turismo Presencial", ("Turismo", "Presencial")),
    "TURISMO EN LINEA": ("Turismo en Línea", ("Turismo", "En línea")),
}


def normalize(s):
    """minusculas, sin acentos/apostrofes, solo alfanumerico + espacios, espacios colapsados."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return s.strip()


def load_descripciones():
    """{(carrera, modalidad): {codigo_CE/CT: descripcion}} desde Cuadros_oficiales_por_carrera.xlsx."""
    wb = load_workbook(CUADROS, data_only=True)
    sheet_to_program = {v[0]: v[1] for v in SHEET_MAP.values()}
    result = {}
    for sheet_name, program in sheet_to_program.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        descs = {}
        for row in ws.iter_rows():
            if len(row) < 10:
                continue
            code, desc = row[0].value, row[9].value
            if isinstance(code, str) and code in COMP_ORDER and isinstance(desc, str) and desc.strip():
                descs.setdefault(code, desc.strip())
        result[program] = descs
    return result


def find_header_row(ws):
    for r in range(1, 20):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().upper() == "ASIGNATURA":
            return r
    raise ValueError(f"No se encontro fila de encabezado ASIGNATURA en hoja {ws.title!r}")


def parse_sheet(ws, descripciones_programa):
    header_row = find_header_row(ws)
    code_row = header_row + 1
    desc_row = header_row + 2
    data_start = header_row + 3

    columns = []  # [{col, label, desc_norm, desc_raw}]
    c = 3
    while True:
        label = ws.cell(code_row, c).value
        if label is None or str(label).strip() == "":
            break
        desc_raw = ws.cell(desc_row, c).value
        desc_norm = normalize(desc_raw)
        if desc_norm:
            columns.append({"col": c, "label": str(label).strip(), "desc_norm": desc_norm, "desc_raw": str(desc_raw).strip()})
        c += 1

    desc_norm_by_code = {code: normalize(desc) for code, desc in descripciones_programa.items()}
    used_codes = set()
    for col in columns:
        match = None
        for code, cnorm in desc_norm_by_code.items():
            if code in used_codes or not cnorm:
                continue
            if col["desc_norm"] == cnorm or col["desc_norm"] in cnorm or cnorm in col["desc_norm"]:
                match = code
                break
        col["codigo"] = match
        if match:
            used_codes.add(match)

    unmatched = [col for col in columns if col["codigo"] is None]
    unused_codes = [code for code in desc_norm_by_code if code not in used_codes]
    if len(unmatched) == 1 and len(unused_codes) == 1:
        unmatched[0]["codigo"] = unused_codes[0]

    competencias = []
    for col in columns:
        # si hay codigo CE/CT, usar la descripcion oficial de Cuadros_oficiales_por_carrera.xlsx
        # (la misma que ya se muestra en el resto del dashboard) en vez del texto crudo de la
        # matriz de articulacion, que a veces trae texto pegado por error de otra carrera.
        descripcion = descripciones_programa.get(col["codigo"], col["desc_raw"]) if col["codigo"] else col["desc_raw"]
        competencias.append({
            "codigo": col["codigo"],
            "label": col["codigo"] or col["label"],
            "descripcion": descripcion,
            "asignaturas": [],
        })

    blank_streak = 0
    r = data_start
    while blank_streak < 20:
        nombre = ws.cell(r, 1).value
        if nombre is None or str(nombre).strip() == "":
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        resultado = ws.cell(r, 2).value
        resultado = str(resultado).strip() if resultado else ""
        for col, comp in zip(columns, competencias):
            mark = ws.cell(r, col["col"]).value
            if mark is not None and str(mark).strip() != "":
                comp["asignaturas"].append({"nombre": str(nombre).strip(), "resultado": resultado})
        r += 1

    return competencias


def main():
    wbA = load_workbook(XLS, data_only=True)
    descripciones = load_descripciones()

    data = {}
    for sheet_name, (cuadros_sheet, program) in SHEET_MAP.items():
        if sheet_name not in wbA.sheetnames:
            print(f"Aviso: hoja {sheet_name!r} no encontrada en {XLS.name}, se omite.")
            continue
        ws = wbA[sheet_name]
        descripciones_programa = descripciones.get(program, {})
        competencias = parse_sheet(ws, descripciones_programa)
        key = program[0] + "|" + program[1]
        data[key] = competencias

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("const ARTICULACION = ")
        json.dump(data, f, ensure_ascii=False)
        f.write(";\n")

    print(f"Escrito {OUT} — {len(data)} programas, {sum(len(v) for v in data.values())} competencias.")


if __name__ == "__main__":
    main()
