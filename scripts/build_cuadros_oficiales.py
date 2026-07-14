"""
Regenera data/Cuadros_oficiales_por_carrera.xlsx a partir de
data/informe_test_aprendizaje.xlsx, replicando el formato original
(bandas de color, encabezados, ítems críticos en rojo).

Uso:
    python scripts/build_cuadros_oficiales.py

Requiere: pandas, openpyxl
"""
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def r1(x):
    """Redondea a 1 decimal con 'mitad hacia arriba' (como Excel), evitando
    el error de punto flotante de round() built-in (ej. round(64.35, 1) == 64.3
    porque 64.35 no es representable exacto en binario)."""
    return float(Decimal(str(float(x))).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


ROOT = Path(__file__).resolve().parent.parent
XLS = ROOT / "data" / "informe_test_aprendizaje.xlsx"
OUT = ROOT / "data" / "Cuadros_oficiales_por_carrera.xlsx"

NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
ZEBRA = "F2F2F2"
RED = "C00000"
GRAY = "595959"

FONT_TITLE = Font(name="Arial", bold=True, size=13, color=NAVY)
FONT_SUBTITLE = Font(name="Arial", size=9, color=GRAY)
FONT_BANNER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_SECTION = Font(name="Arial", bold=True, size=10, color="000000")
FONT_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10, color="000000")
FONT_BODY_BOLD = Font(name="Arial", bold=True, size=10, color="000000")
FONT_CRITICAL = Font(name="Arial", bold=True, size=10, color=RED)

FILL_BANNER = PatternFill("solid", fgColor=NAVY)
FILL_SECTION = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_ZEBRA = PatternFill("solid", fgColor=ZEBRA)
FILL_NONE = PatternFill(fill_type=None)

CENTER = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")

# (sheet_name, carrera_nombre en informe, modalidad_nombre en informe, texto para el titulo)
PROGRAMS = [
    ("Administración", "Administración", "Presencial", "Administración (Presencial)"),
    ("Contabilidad", "Contabilidad y Auditoría", "Presencial", "Contabilidad y Auditoría (Presencial)"),
    ("Economía Presencial", "Economía", "Presencial", "Economía (Presencial)"),
    ("Economía en Línea", "Economía", "En línea", "Economía (En línea)"),
    ("Turismo Presencial", "Turismo", "Presencial", "Turismo (Presencial)"),
    ("Turismo en Línea", "Turismo", "En línea", "Turismo (En línea)"),
]

COMP_ORDER = ["CE1", "CE2", "CE3", "CE4", "CT1", "CT2", "CT3", "CT4"]


def set_row(ws, row, values, font, fill, align=CENTER, first_align=None):
    for i, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = first_align if (i == 1 and first_align) else align


def banner(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_BANNER
    cell.fill = FILL_BANNER
    cell.alignment = LEFT
    for c in range(2, 10):
        ws.cell(row=row, column=c).fill = FILL_BANNER


def section(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = LEFT
    for c in range(2, 10):
        ws.cell(row=row, column=c).fill = FILL_SECTION


def main():
    cg = pd.read_excel(XLS, sheet_name="Cohorte_Global")
    cc = pd.read_excel(XLS, sheet_name="Cohorte_x_Competencia")
    ai = pd.read_excel(XLS, sheet_name="Aciertos_x_Item")
    eg = pd.read_excel(XLS, sheet_name="Estudiante_Global")
    ec = pd.read_excel(XLS, sheet_name="Estudiante_x_Competencia")
    nitems = ec.groupby(["nombre_test", "competencia"])["n_items"].first().to_dict()

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, carrera, modalidad, titulo_prog in PROGRAMS:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = True
        for col, width in {"A": 13, "C": 9, "D": 14, "E": 17, "F": 10, "I": 12}.items():
            ws.column_dimensions[col].width = width

        r = 1
        ws.cell(row=r, column=1, value=f"Cuadros oficiales verificados — {titulo_prog}").font = FONT_TITLE
        r += 1
        ws.cell(row=r, column=1, value=(
            "Valores oficiales (informe_test_aprendizaje.xlsx), recalculados y verificados. "
            "Escala: I 0–49, ED 50–69, S 70–89, SO 90–100. Listos para el informe."
        )).font = FONT_SUBTITLE
        r += 2

        courses = cg[(cg["carrera_nombre"] == carrera) & (cg["modalidad_nombre"] == modalidad)].sort_values("nro_test")
        if courses.empty:
            print(f"AVISO: sin datos para {titulo_prog}, hoja creada vacia.")
            continue

        for _, row in courses.iterrows():
            nt, ta = row["nombre_test"], int(row["nro_test"])
            n_total = int(row["n_estudiantes_unicos"])

            banner(ws, r, f"══ TA{ta} — {n_total} estudiantes ══")
            r += 1
            section(ws, r, "Resultado global de cohorte")
            r += 1
            set_row(ws, r, ["Promedio %", "Nivel", "Mediana %", "Mínimo %", "Máximo %", "Desv.",
                             "% Insuf.", "% En des.", "% Satisf./Sobres."], FONT_HEADER, FILL_HEADER)
            r += 1
            pct_sat, pct_sob = r1(row["pct_satisfactorio"]), r1(row["pct_sobresaliente"])
            set_row(ws, r, [
                r1(row["promedio_global"]), row["nivel_global"], r1(row["mediana_global"]),
                r1(row["minimo_global"]), r1(row["maximo_global"]), r1(row["sd_global"]),
                r1(row["pct_insuficiente"]), r1(row["pct_en_desarrollo"]), f"{pct_sat} / {pct_sob}",
            ], FONT_BODY, FILL_NONE)
            r += 1

            counts = eg[eg["nombre_test"] == nt]["nivel_global"].value_counts().to_dict()
            def cnt(prefix):
                for k, v in counts.items():
                    if k.startswith(prefix):
                        return int(v)
                return 0
            ws.cell(row=r, column=1, value="Distribución (N° estud.):").font = FONT_BODY_BOLD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            for col, txt in zip([4, 5, 6, 7], [f"Insuf: {cnt('Insuf')}", f"En des: {cnt('En des')}",
                                                 f"Satisf: {cnt('Satisf')}", f"Sobres: {cnt('Sobres')}"]):
                cell = ws.cell(row=r, column=col, value=txt)
                cell.font = FONT_BODY
                cell.alignment = CENTER
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
            tot_cell = ws.cell(row=r, column=8, value=f"Total: {n_total}")
            tot_cell.font = FONT_BODY
            tot_cell.alignment = CENTER
            r += 2

            section(ws, r, "Logro y distribución por competencia")
            r += 1
            set_row(ws, r, ["Código", "Tipo", "N° preg.", "Prom. logro %", "Nivel",
                             "% Insuf.", "% En des.", "% Satisf.", "% Sobres."], FONT_HEADER, FILL_HEADER)
            r += 1
            comp_rows = cc[(cc["nombre_test"] == nt)]
            comp_rows = comp_rows.set_index("competencia").reindex(
                [c for c in COMP_ORDER if c in comp_rows["competencia"].values]
            ).reset_index()
            for i, crow in comp_rows.iterrows():
                fill = FILL_ZEBRA if i % 2 == 1 else FILL_NONE
                code = crow["competencia"]
                tipo = "Específica" if code.startswith("CE") else "Transversal"
                n_it = int(nitems.get((nt, code), 0))
                set_row(ws, r, [
                    code, tipo, n_it, r1(crow["promedio_logro"]), crow["nivel_cohorte"],
                    r1(crow["pct_insuficiente"]), r1(crow["pct_en_desarrollo"]),
                    r1(crow["pct_satisfactorio"]), r1(crow["pct_sobresaliente"]),
                ], FONT_BODY, fill)
                ws.cell(row=r, column=1).font = FONT_BODY_BOLD
                r += 1
            r += 1

            section(ws, r, "Ítems críticos (<50% de aciertos, nivel Insuficiente)")
            r += 1
            crit = ai[(ai["nombre_test"] == nt) & (ai["pct_aciertos"] < 50)].sort_values("pct_aciertos")
            set_row(ws, r, ["Código", "Competencias", "", "", "", "", "", "", "% Aciertos"], FONT_HEADER, FILL_HEADER)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            r += 1
            if crit.empty:
                ws.cell(row=r, column=1, value="Ninguno (ningún ítem por debajo del 50%).").font = FONT_BODY
                r += 1
            else:
                for i, (_, irow) in enumerate(crit.iterrows()):
                    fill = FILL_ZEBRA if i % 2 == 1 else FILL_NONE
                    ws.cell(row=r, column=1, value=irow["codigo_item"]).font = FONT_BODY
                    ws.cell(row=r, column=1).fill = fill
                    ws.cell(row=r, column=1).alignment = CENTER
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
                    comp_cell = ws.cell(row=r, column=2, value=irow["competencias_evaluadas"])
                    comp_cell.font = FONT_BODY
                    comp_cell.fill = fill
                    comp_cell.alignment = LEFT
                    for c in range(3, 9):
                        ws.cell(row=r, column=c).fill = fill
                    val_cell = ws.cell(row=r, column=9, value=r1(irow["pct_aciertos"]))
                    val_cell.font = FONT_CRITICAL
                    val_cell.fill = fill
                    val_cell.alignment = CENTER
                    r += 1
            r += 2

    wb.save(OUT)
    print(f"Escrito {OUT} con {len(wb.sheetnames)} hojas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
