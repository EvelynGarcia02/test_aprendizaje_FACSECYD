"""
Regenera js/data.js a partir de data/informe_test_aprendizaje.xlsx.

Uso:
    python scripts/extract_data.py

Requiere: pandas, openpyxl  (pip install pandas openpyxl)

Lee las hojas Cohorte_Global, Cohorte_x_Competencia, Aciertos_x_Item,
Estudiante_Global y Estudiante_x_Competencia de data/informe_test_aprendizaje.xlsx
y escribe un unico archivo JS con "const DATA = {...};" que el dashboard
carga directamente (sin fetch, para que funcione abriendo index.html sin
necesidad de un servidor local).
"""
import json
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLS = ROOT / "data" / "informe_test_aprendizaje.xlsx"
CUADROS = ROOT / "data" / "Cuadros_oficiales_por_carrera.xlsx"
OUT = ROOT / "js" / "data.js"

COMP_ORDER = ["CE1", "CE2", "CE3", "CE4", "CT1", "CT2", "CT3", "CT4"]

# hoja de Cuadros_oficiales_por_carrera.xlsx -> (carrera, modalidad) tal como
# aparecen en informe_test_aprendizaje.xlsx
SHEET_TO_PROGRAM = {
    "Administración": ("Administración", "Presencial"),
    "Contabilidad": ("Contabilidad y Auditoría", "Presencial"),
    "Economía Presencial": ("Economía", "Presencial"),
    "Economía en Línea": ("Economía", "En línea"),
    "Turismo Presencial": ("Turismo", "Presencial"),
    "Turismo en Línea": ("Turismo", "En línea"),
}


def load_descripciones():
    """Lee la columna 'Descripción' (col. J) de cada hoja de
    Cuadros_oficiales_por_carrera.xlsx -> {(carrera, modalidad): {competencia: descripcion}}.
    Esa columna se agregó a mano en ese libro (build_cuadros_oficiales.py no la genera),
    por eso se lee directo de celdas en vez de regenerar el archivo."""
    if not CUADROS.exists():
        return {}
    wb = load_workbook(CUADROS, data_only=True)
    result = {}
    for sheet_name, program in SHEET_TO_PROGRAM.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        descs = {}
        for row in ws.iter_rows():
            if len(row) < 10:
                continue
            code, desc = row[0].value, row[9].value
            if isinstance(code, str) and code in COMP_ORDER and isinstance(desc, str) and desc.strip():
                descs.setdefault(code, desc.strip())
        result[program] = descs
    return result


def r1(x):
    """Redondea a 1 decimal con 'mitad hacia arriba' (como Excel), evitando
    el error de punto flotante de round() built-in (ej. round(64.35, 1) == 64.3
    porque 64.35 no es representable exacto en binario)."""
    return float(Decimal(str(float(x))).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


def main():
    cg = pd.read_excel(XLS, sheet_name="Cohorte_Global")
    cc = pd.read_excel(XLS, sheet_name="Cohorte_x_Competencia")
    ai = pd.read_excel(XLS, sheet_name="Aciertos_x_Item")
    eg = pd.read_excel(XLS, sheet_name="Estudiante_Global")
    ec = pd.read_excel(XLS, sheet_name="Estudiante_x_Competencia")

    nitems = ec.groupby(["nombre_test", "competencia"])["n_items"].first().to_dict()
    curso_a_programa = {r["nombre_test"]: (r["carrera_nombre"], r["modalidad_nombre"]) for _, r in cg.iterrows()}
    descripciones = load_descripciones()

    courses = []
    for _, r in cg.iterrows():
        nt = r["nombre_test"]
        counts = eg[eg["nombre_test"] == nt]["nivel_global"].value_counts().to_dict()

        def cnt(prefix):
            for k, v in counts.items():
                if k.startswith(prefix):
                    return int(v)
            return 0

        courses.append({
            "id": nt,
            "carrera": r["carrera_nombre"],
            "modalidad": r["modalidad_nombre"],
            "ta": int(r["nro_test"]),
            "n": int(r["n_estudiantes_unicos"]),
            "prom": r1(r["promedio_global"]),
            "mediana": r1(r["mediana_global"]),
            "min": r1(r["minimo_global"]),
            "max": r1(r["maximo_global"]),
            "sd": r1(r["sd_global"]),
            "nivel": r["nivel_global"],
            "pct": {
                "insuf": r1(r["pct_insuficiente"]),
                "ed": r1(r["pct_en_desarrollo"]),
                "sat": r1(r["pct_satisfactorio"]),
                "sob": r1(r["pct_sobresaliente"]),
            },
            "counts": {
                "insuf": cnt("Insuf"),
                "ed": cnt("En des"),
                "sat": cnt("Satisf"),
                "sob": cnt("Sobres"),
            },
        })

    competencias = []
    for _, r in cc.iterrows():
        key = (r["nombre_test"], r["competencia"])
        programa = curso_a_programa.get(r["nombre_test"])
        descripcion = descripciones.get(programa, {}).get(r["competencia"], "")
        competencias.append({
            "curso_id": r["nombre_test"],
            "competencia": r["competencia"],
            "tipo": "Específica" if r["competencia"].startswith("CE") else "Transversal",
            "n_items": int(nitems.get(key, 0)),
            "prom": r1(r["promedio_logro"]),
            "nivel": r["nivel_cohorte"],
            "descripcion": descripcion,
            "pct": {
                "insuf": r1(r["pct_insuficiente"]),
                "ed": r1(r["pct_en_desarrollo"]),
                "sat": r1(r["pct_satisfactorio"]),
                "sob": r1(r["pct_sobresaliente"]),
            },
        })

    items = []
    for _, r in ai.iterrows():
        items.append({
            "curso_id": r["nombre_test"],
            "codigo": r["codigo_item"],
            "pct": r1(r["pct_aciertos"]),
            "competencias": r["competencias_evaluadas"],
        })

    data = {"courses": courses, "competencias": competencias, "items": items}

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("const DATA = ")
        json.dump(data, f, ensure_ascii=False)
        f.write(";\n")

    print(f"Escrito {OUT} — {len(courses)} aplicaciones del test, {len(competencias)} filas de competencia, {len(items)} items.")


if __name__ == "__main__":
    main()
